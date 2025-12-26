"""
Data Service - Handle Excel file processing and JSON conversion
Ported from reference implementation
"""

import json
import os
import re
import uuid
from datetime import datetime
from typing import Dict, Any, List, Optional
import pandas as pd
import openpyxl
from pathlib import Path

from ..core.logger import logger

try:
    import orjson
    HAS_ORJSON = True
except ImportError:
    HAS_ORJSON = False

# Custom Exceptions
class DataServiceError(Exception):
    """Custom exception for data service errors"""
    def __init__(self, message: str):
        super().__init__(message)
        self.message = message

class DataService:
    """Service for processing Excel files and converting to JSON"""

    def __init__(self, upload_dir: str = "data/uploads"):
        """Initialize data service with upload directory"""
        self.upload_dir = Path(upload_dir)
        self.upload_dir.mkdir(parents=True, exist_ok=True)
        self.index_file = self.upload_dir / "upload_index.json"
        self.upload_index = {}
        self._load_index()

    def _load_index(self):
        """Load upload index from disk, rebuild if not exists or corrupted"""
        if self.index_file.exists():
            try:
                with open(self.index_file, 'r', encoding='utf-8') as f:
                    self.upload_index = json.load(f)
                logger.info(f"Loaded {len(self.upload_index)} files from index")
                
                # If index is empty but files exist, rebuild
                if not self.upload_index and self._has_excel_files():
                    logger.warning("Index is empty but Excel files exist, rebuilding...")
                    self._rebuild_index()
            except Exception as e:
                logger.warning(f"Failed to load index: {e}, rebuilding from scratch")
                self._rebuild_index()
        else:
            # No index file - scan and build initial index
            if self._has_excel_files():
                logger.info("Index file not found but Excel files exist, building initial index...")
                self._rebuild_index()
            else:
                logger.info("No index file and no Excel files, starting fresh")
                self.upload_index = {}
    
    def _has_excel_files(self) -> bool:
        """Quick check if any Excel files exist in upload directory"""
        if not self.upload_dir.exists():
            return False
        
        try:
            for year_entry in os.scandir(self.upload_dir):
                if year_entry.is_dir():
                    for month_entry in os.scandir(year_entry.path):
                        if month_entry.is_dir():
                            for day_entry in os.scandir(month_entry.path):
                                if day_entry.is_dir():
                                    for file_entry in os.scandir(day_entry.path):
                                        if file_entry.is_file() and file_entry.name.endswith(('.xlsx', '.xls')):
                                            return True
            return False
        except Exception:
            return False

    def _rebuild_index(self):
        """Rebuild index by scanning directories"""
        self.upload_index = {}
        if not self.upload_dir.exists():
            return

        try:
            for root, _, files in os.walk(self.upload_dir):
                for file in files:
                    if file.endswith(('.xlsx', '.xls')):
                        # Extract original filename from saved filename
                        # Format: date_time_original.xlsx
                        parts = file.split('_', 2)
                        if len(parts) >= 3:
                            original_name = parts[2]
                            self.upload_index[original_name] = str(Path(root) / file)
            self._save_index()
            logger.info(f"Rebuilt index with {len(self.upload_index)} files")
        except Exception as e:
            logger.error(f"Error rebuilding index: {e}")
    
    def _save_index(self):
        """Save upload index to disk"""
        try:
            with open(self.index_file, 'w', encoding='utf-8') as f:
                json.dump(self.upload_index, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Failed to save index: {e}")
    
    def process_excel_upload(self, file_content: bytes, filename: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Main entry point: Upload Excel → Convert to JSON
        """
        try:
            # Step 1: Check existing file
            existing_file = self.upload_index.get(filename)
            if existing_file and Path(existing_file).exists():
                logger.info(f"File '{filename}' already exists, using existing version")
                excel_file_path = existing_file
            else:
                # Save new Excel file
                now = datetime.now()
                year_folder = now.strftime("%Y")
                month_folder = now.strftime("%m")
                day_folder = now.strftime("%d")
                
                date_path = self.upload_dir / year_folder / month_folder / day_folder
                date_path.mkdir(parents=True, exist_ok=True)
                
                date_part = now.strftime("%d%m%y")
                time_part = now.strftime("%H%M%S")
                saved_filename = f"{date_part}_{time_part}_{filename}"
                excel_file_path = str(date_path / saved_filename)
                
                with open(excel_file_path, 'wb') as f:
                    f.write(file_content)
                
                self.upload_index[filename] = excel_file_path
                self._save_index()
                logger.info(f"Saved new file: {excel_file_path}")
            
            # Step 2: Read Excel and find header
            wb = openpyxl.load_workbook(excel_file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            
            # Find target sheet (case-insensitive)
            target_sheet_index = 0  # Default: first sheet
            target_sheet_name = sheet_names[0]  # Default sheet name
            
            if sheet_name:
                # Case-insensitive search
                sheet_name_lower = sheet_name.strip().lower()
                found = False
                
                for idx, name in enumerate(sheet_names):
                    if name.lower() == sheet_name_lower:
                        target_sheet_index = idx
                        target_sheet_name = name
                        found = True
                        logger.info(f"Found sheet '{name}' (requested: '{sheet_name}')")
                        break
                
                if not found:
                    wb.close()
                    raise DataServiceError(
                        f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(sheet_names)}"
                    )
            else:
                logger.info(f"No sheet specified, using first sheet: '{target_sheet_name}'")
            
            # Get worksheet
            ws = wb[target_sheet_name]
            
            # Find row with 'STT' header
            header_row_idx = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if row and any(cell and str(cell).strip().upper() == "STT" for cell in row):
                    header_row_idx = i
                    logger.info(f"Found 'STT' header at row {i}")
                    break
            wb.close()
            
            # Read Excel with correct header and sheet
            df = pd.read_excel(excel_file_path, sheet_name=target_sheet_name, header=header_row_idx)
            
            # Remove STT column and empty rows
            if 'STT' in df.columns:
                df = df.drop('STT', axis=1)
            df = df.dropna(how='all')
            
            # Remove unnamed columns
            unnamed_cols = [col for col in df.columns if str(col).startswith('Unnamed:')]
            if unnamed_cols:
                df = df.drop(columns=unnamed_cols)
            
            logger.info(f"Loaded {len(df)} rows, {len(df.columns)} columns")
            
            # Step 3: Detect column mapping
            column_mapping = self._detect_columns(df.columns)
            logger.info(f"Detected columns: {column_mapping}")
            
            # Step 4: Process each row
            batch_id = str(uuid.uuid4())
            processed_records = []
            
            for idx, row in df.iterrows():
                record = {}
                
                # Process each column
                for original_col in df.columns:
                    standard_field = column_mapping.get(original_col, original_col)
                    value = row[original_col]

                    # Convert value based on field type
                    record[standard_field] = self._convert_field_value(standard_field, value)

                # Special handling: Extract rooms from "loai_phong" like "2PN + 1 WC"
                room_fields_to_check = ['so_phong_ngu', 'loai_can']
                for field in room_fields_to_check:
                    if field in record and record[field]:
                        phong_ngu, phong_wc = self._extract_rooms_from_loai_phong(record[field])
                        if phong_ngu is not None:
                            record['so_phong_ngu'] = phong_ngu
                        if phong_wc is not None and ('so_phong_wc' not in record or not record['so_phong_wc']):
                            record['so_phong_wc'] = phong_wc
                        # Remove the original field if it's been processed
                        if field != 'so_phong_ngu':
                            record.pop(field, None)
                        break
                
                # Extract toa/tang/can from ma_can if missing
                if 'ma_can' in record and record['ma_can']:
                    known_toa = record.get('toa') if 'toa' in record and record['toa'] else None
                    ma_can_data = self._extract_from_ma_can(record['ma_can'], known_toa)

                    if 'toa' not in record or not record['toa']:
                        record['toa'] = ma_can_data.get('toa')
                    if 'tang' not in record or not record['tang']:
                        record['tang'] = ma_can_data.get('tang')
                    if 'can' not in record or not record['can']:
                        record['can'] = ma_can_data.get('can')
                
                # Add metadata
                record['_id'] = f"{batch_id}_{idx}"
                record['_source_file'] = filename
                record['_processed_at'] = datetime.now().isoformat()
                
                processed_records.append(record)
            
            logger.info(f"Processed {len(processed_records)} records")
            
            # Step 5: Save JSON file
            excel_path = Path(excel_file_path)
            json_folder = excel_path.parent
            
            excel_filename = excel_path.name
            parts = excel_filename.split('_', 2)
            if len(parts) >= 3:
                date_part = parts[0]
                time_part = parts[1]
                original_name = parts[2].replace('.xlsx', '').replace('.xls', '').lower()
            else:
                now = datetime.now()
                date_part = now.strftime("%d%m%y")
                time_part = now.strftime("%H%M%S")
                original_name = excel_filename.replace('.xlsx', '').replace('.xls', '').lower()
            
            json_filename = f"{date_part}_{time_part}_{original_name}.json"
            json_path = json_folder / json_filename
            
            json_data = {
                "batch_id": batch_id,
                "source_file": filename,
                "processed_at": datetime.now().isoformat(),
                "total_records": len(processed_records),
                "excel_info": {
                    "sheets": sheet_names,
                    "selected_sheet": target_sheet_name,
                    "header_row": header_row_idx,
                    "detected_columns": column_mapping
                },
                "data": processed_records
            }
            
            # Save JSON
            if HAS_ORJSON:
                with open(json_path, 'wb') as f:
                    f.write(orjson.dumps(json_data, option=orjson.OPT_INDENT_2))
            else:
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(json_data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"Saved JSON: {json_path}")
            
            return {
                "message": f"Successfully processed {filename}",
                "batch_id": batch_id,
                "total_records": len(processed_records),
                "json_file": str(json_path),
                "preview": processed_records[:3],
                "error": None
            }
            
        except Exception as e:
            logger.error(f"Error processing Excel: {e}", exc_info=True)
            return {
                "message": "Failed to process Excel file",
                "batch_id": None,
                "total_records": 0,
                "json_file": None,
                "preview": [],
                "error": str(e)
            }
    
    def _detect_columns(self, columns: List[str]) -> Dict[str, str]:
        """Detect column mapping"""
        keywords = {
            'ma_can': ['mã căn', 'ma can', 'mã', 'apartment code'],
            'toa': ['tòa', 'toa', 'tower', 'building'],
            'tang': ['tầng', 'tang', 'floor'],
            'can': ['căn', 'can', 'unit'],
            'loai_can': ['loại căn', 'loại hình', 'loại pn' ,'loại phòng', 'loai phong', 'loại căn hộ'],
            'so_phong_ngu': ['phòng ngủ', 'phong ngu', 'bedroom', 'pn'],
            'so_phong_wc': ['phòng wc', 'phong wc', 'wc', 'toilet', 'bathroom'],
            'khach_hang': ['khách hàng', 'khach hang', 'customer', 'owner', 'đl nắm căn'],
            'dien_thoai': ['điện thoại', 'dien thoai', 'phone', 'sdt'],
            'mang_vien_thong': ['mạng viễn thông', 'mang vien thong', 'mạng', 'mang', 'network'],
            'email': ['email', 'e-mail'],
            'huong': ['hướng', 'huong', 'direction', 'hướng ban công'],
            'view': ['view', 'tầm nhìn', 'tam nhin', 'tầm view', 'tầm nhìn'],
            'the_gio': ['thế gió', 'the gio'],
            'dien_tich': ['diện tích', 'dien tich', 'area', 'dt', 'm2', 'diện tích thông thuỷ', 'DT thông thuỷ', 'DT thông thủy', 'DIỆN TÍCH THÔNG THỦY','DT','Dien tich', 'dien tich thong thuy', 'dien tich thong thuỷ', "DTTT"],
            'dien_tich_tim_tuong': ['DT tim tường', 'DIỆN TÍCH TIM TƯỜNG', 'dt tim tường', 'dien tich tim tuong', 'diện tích tim tường', 'DT TIM TƯỜNG', 'DIỆN TÍCH TIM TƯỜNG', 'DTTT', 'Dien tich tim tuong', 'dien tich tim tuong', 'DT tim tuong'],
            'ngay_update': ['ngày update', 'ngay update', 'update', 'cập nhật'],
            'nhu_cau': ['nhu cầu', 'nhu cau', 'demand', 'b/t', 'status'],
            'noi_that': ['nội thất', 'noi that', 'furniture'],
            'gia_net': ['giá net', 'gia net', 'net price', 'giá trước VAT + KPBT', 'giá chưa VAT + KPBT', 'giá chưa VAT và KPBT'],
            'gia_ban': ['giá bán', 'gia ban', 'sale price', 'selling', 'tổng giá', 'giá tiêu chuẩn', 'giá trần', 'tổng giá bán niêm yết', 'giá bán niêm yết', 'giá tiêu chuẩn bàn giao cơ bản', 'tcbg', 'TCBG', 'dự kiến giá', 'du kien gia'],
            'gia_bao_khach': ['giá báo khách', 'gia bao khach', 'quoted'],
            'gia_thue': ['giá thuê', 'gia thue', 'rental', 'rent'],
            'ghi_chu': ['ghi chú', 'ghi chu', 'note', 'remark', 'pass', 'comment', 'lưu ý', 'luu y'],
            'du_an': ['dự án', 'du an', 'project'],
            'phan_khu': ['phân khu', 'phan khu', 'phase', 'khu', 'địa chỉ', 'dia chi'],
            'chi_ranh': ['chi ranh', 'zone', 'vùng'],
            'ngay_bat_dau': ['ngày bắt đầu', 'ngay bat dau', 'start date'],
            'ngay_ket_thuc': ['ngày kết thúc', 'ngay ket thuc', 'end date'],
            'ngay_vao_coc': ['ngày vào cọc', 'ngay vao coc', 'deposit date'],
            'dai_ly': ['đại lý', 'dai ly', 'agent', 'agency', 'pkd/đại lý', 'tcm', 'ak', 'cdt', 'dtl', 'minh tam land'],
            'tinh_trang': ['tình trạng', 'tinh trang', 'status', 'condition', 'tình trạng ký chuyển nhượng', 'nguồn', 'csbh'],
            'link_ptg': ['phiếu tính giá', 'ptg', 'link', 'drive', 'phiếu tính giá và hình vẽ căn'],
            'link_info': ['info & ptg', 'info', 'thông tin'],
        }
        
        mapping = {}
        for col in columns:
            col_lower = str(col).strip().lower()
            matched_field = None
            priority_order = ['dien_tich_tim_tuong', 'dien_tich', 'so_phong_ngu']

            for field_name in priority_order:
                if field_name in keywords:
                    if any(kw in col_lower for kw in keywords[field_name]):
                        matched_field = field_name
                        break

            if not matched_field:
                for field_name, field_keywords in keywords.items():
                    if field_name not in priority_order:
                        if any(kw in col_lower for kw in field_keywords):
                            matched_field = field_name
                            break

            mapping[col] = matched_field if matched_field else col
        
        return mapping
    
    def _extract_from_ma_can(self, ma_can: str, known_toa: Optional[str] = None) -> Dict[str, Any]:
        """Extract toa/tang/can from ma_can"""
        if not ma_can or pd.isna(ma_can):
            return {}

        ma_can_clean = str(ma_can).strip().replace(' ', '')

        if known_toa and ma_can_clean.startswith(known_toa):
            remaining = ma_can_clean[len(known_toa):]  
            parts = remaining.split('.')
            if len(parts) >= 2:
                result = {'toa': known_toa}
                floor_part = '.'.join(parts[:-1])  
                unit_part = parts[-1]  

                if '.' in floor_part:
                    floor_digits = floor_part.split('.')[-1]  
                else:
                    floor_digits = floor_part

                if len(floor_digits) >= 2:
                    try:
                        floor_num = floor_digits[-2:] if len(floor_digits) > 3 else floor_digits
                        result['tang'] = int(floor_num)
                    except:
                        result['tang'] = floor_digits

                try:
                    result['can'] = int(unit_part)
                except:
                    result['can'] = unit_part

                return result

        parts = re.split(r'[\.\-_/]+', ma_can_clean)
        if len(parts) < 2:
            return {}

        result = {}
        result['toa'] = parts[0]

        if len(parts) >= 2:
            try:
                result['tang'] = int(parts[1])
            except:
                result['tang'] = parts[1]

        if len(parts) >= 3:
            try:
                result['can'] = int(parts[2])
            except:
                result['can'] = parts[2]

        return result

    def _extract_rooms_from_loai_phong(self, loai_phong: str) -> tuple[Optional[int], Optional[int]]:
        """Extract so_phong_ngu and so_phong_wc"""
        if not loai_phong or pd.isna(loai_phong):
            return None, None

        loai_phong_str = str(loai_phong).strip().upper()
        if loai_phong_str == "1BR":
            return 1, None

        phong_ngu = None
        phong_wc = None

        pn_match = re.search(r'(\d+)\s*PN', loai_phong_str)
        if pn_match:
            phong_ngu = int(pn_match.group(1))

        wc_match = re.search(r'(\d+)\s*WC', loai_phong_str)
        if wc_match:
            phong_wc = int(wc_match.group(1))

        if phong_wc is None:
            plus_match = re.search(r'PN\s*\+\s*(\d+)', loai_phong_str)
            if plus_match:
                phong_wc = int(plus_match.group(1))

        return phong_ngu, phong_wc

    def _convert_field_value(self, field_name: str, value: Any) -> Any:
        """Convert field value based on field type"""
        if value is None or pd.isna(value):
            return None
        
        if isinstance(value, str):
            value_clean = value.strip().lower()
            if value_clean in ['', '-', '--', '—', 'n/a', 'na', 'null', 'none']:
                return None
        
        if field_name in ['gia_net', 'gia_ban', 'gia_bao_khach', 'gia_thue']:
            return self._convert_price(value)
        
        if field_name in ['dien_tich', 'dien_tich_tim_tuong', 'dien_tich_thong_thuy']:
            return self._convert_area(value)

        if field_name in ['so_phong_ngu', 'so_phong_wc', 'tang', 'can']:
            return self._convert_int(value)

        if field_name in ['ngay_bat_dau', 'ngay_ket_thuc', 'ngay_vao_coc', 'ngay_update']:
            return self._convert_date(value)

        return self._convert_text(value)
    
    def _convert_price(self, value: Any) -> Any:
        """Convert price to VND"""
        # Helper to process single number
        def process_number(num):
            # If value is very large (> 100 million), assume it's already VND
            if num > 100_000_000:
                return int(num)
            # If value is small (e.g. 3.5, 10, 3950), assume it's in Millions
            # Example: 3950 -> 3,950,000,000 (3.95 billion)
            # Example: 3.5 -> 3,500,000 (3.5 million? Or 3.5 billion?)
            # Context matters. But usually real estate prices are > 500 million.
            # If num < 100: likely Billion (e.g. 3.5 tỷ)
            # If num >= 100: likely Million (e.g. 3950 triệu)
            
            if num < 100: 
                # Ambiguous: 3.5 could be 3.5 Million (rent) or 3.5 Billion (sale)
                # But for 'gia_ban', usually Billion.
                # Let's assume Billion for very small numbers if it's a sale price.
                return int(num * 1_000_000_000)
            else:
                # Assume Million
                return int(num * 1_000_000)

        if isinstance(value, (int, float)):
            return process_number(float(value))
        
        if not isinstance(value, str):
            return None
        
        value_str = value.strip().lower()
        if not value_str: return None

        # Check for explicit units
        multiplier = 1_000_000 # Default to Million
        if 'tỷ' in value_str or 'ty' in value_str:
            multiplier = 1_000_000_000
        elif 'tr' in value_str or 'triệu' in value_str or 'trieu' in value_str:
            multiplier = 1_000_000
        elif 'nghìn' in value_str or 'nghin' in value_str or 'k' in value_str:
            multiplier = 1_000
        
        # Extract numbers
        # Handle formats like "3,950" or "3.950" -> need to be careful with separators
        # Vietnamese often use '.' for thousands and ',' for decimals, or vice versa.
        # Safest is to remove non-numeric chars except one separator if present.
        
        # Simple extraction for now: find all numbers
        # Replace ',' with '.' if it looks like decimal? 
        # Actually, "3,950" in Excel might be read as string "3,950". 
        # If we just remove ',' and '.' and parse? No, "3.5" becomes 35.
        
        # Let's try to parse the string as a float first after cleaning
        clean_str = value_str.replace('tỷ', '').replace('ty', '').replace('triệu', '').replace('tr', '').replace('vnd', '').replace('đ', '').strip()
        
        # Try to handle "3,95" vs "3.95"
        # If contains both . and , -> remove the one that appears more or first?
        # Standardize: replace ',' with '.' if it's likely a decimal separator
        
        numbers = re.findall(r'\d+(?:[.,]\d+)?', clean_str)
        
        converted = []
        for num_str in numbers:
            try:
                # Normalize separator: replace ',' with '.' for float parsing
                num_val = float(num_str.replace(',', '.'))
                
                if 'tỷ' in value_str or 'ty' in value_str:
                    converted.append(int(num_val * 1_000_000_000))
                elif 'tr' in value_str or 'triệu' in value_str:
                    converted.append(int(num_val * 1_000_000))
                else:
                    # No explicit unit, use heuristic
                    converted.append(process_number(num_val))
            except:
                continue
        
        if not converted:
            return None
        
        if len(converted) == 1:
            return converted[0]
        
        return {
            "min": min(converted),
            "max": max(converted),
            "unit": "VND"
        }
    
    def _convert_area(self, value: Any) -> Optional[float]:
        """Convert area to float"""
        if isinstance(value, (int, float)):
            return round(float(value), 2)
        
        if not isinstance(value, str):
            return None
        
        value_str = value.replace(',', '.').replace(' ', '')
        numbers = re.findall(r'\d+(?:\.\d+)?', value_str)
        
        if numbers:
            try:
                return round(float(numbers[0]), 2)
            except:
                return None
        return None
    
    def _convert_int(self, value: Any) -> Optional[int]:
        """Convert to integer"""
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        if isinstance(value, str):
            numbers = re.findall(r'\d+', value)
            if numbers:
                try:
                    return int(numbers[0])
                except:
                    return None
        return None
    
    def _convert_date(self, value: Any) -> Optional[str]:
        """Convert date to ISO format"""
        if value is None or pd.isna(value):
            return None
        if hasattr(value, 'strftime'):
            return value.strftime('%Y-%m-%d')
        if not isinstance(value, str):
            return None
        
        value_clean = value.strip()
        if not value_clean or value_clean.lower() in ['', '-', '--', '—', 'n/a', 'na', 'null', 'none']:
            return None

        try:
            if '/' in value_clean or '-' in value_clean:
                parts = re.split(r'[/-]', value_clean)
                if len(parts) == 3:
                    try:
                        day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
                        if year < 100: year += 2000
                        if 1 <= day <= 31 and 1 <= month <= 12 and year >= 2000:
                            return f"{year:04d}-{month:02d}-{day:02d}"
                    except:
                        pass
                    try:
                        if len(parts[0]) == 4:
                            year, month, day = int(parts[0]), int(parts[1]), int(parts[2])
                            if 1 <= month <= 12 and 1 <= day <= 31 and year >= 2000:
                                return f"{year:04d}-{month:02d}-{day:02d}"
                    except:
                        pass
            return value_clean
        except:
            return value_clean

    def _convert_text(self, value: Any) -> Optional[str]:
        """Convert to clean text"""
        if isinstance(value, str):
            text = value.strip()
            return text if text else None
        if isinstance(value, (int, float)):
            return str(value)
        return None

# Global instance
data_service = DataService()
